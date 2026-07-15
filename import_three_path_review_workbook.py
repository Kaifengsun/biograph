"""Apply human decisions from the review workbook without changing its source pack."""

from __future__ import annotations

import argparse
import json
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from build_regulatory_evidence_graph import read_jsonl
from three_path_evaluation import review_ledger, sha256_file, split_ids, split_path_nodes


DEFAULT_CORPUS = Path("data/staging/enrichment_full_2026-07-deepseek-v4-pro-v4")
DEFAULT_GRAPH = Path("artifacts/regulatory_evidence_graph/deepseek-v4-pro-v4-build5-regulatory-fda")


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def frozen_chunk_ids(corpus_dir: Path) -> set[str]:
    chunk_ids = set()
    for path in sorted(corpus_dir.glob("*_enriched.json")):
        chunk_ids.update(str(row.get("chunk_id", "")) for row in read_json(path) if str(row.get("chunk_id", "")))
    if not chunk_ids:
        raise ValueError(f"no frozen chunks found in {corpus_dir}")
    return chunk_ids


def graph_node_ids(graph_dir: Path) -> set[str]:
    nodes_path = graph_dir / "nodes.jsonl"
    return {str(row["id"]) for row in read_jsonl(nodes_path)}


def workbook_decisions(path: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Review Queue" not in workbook.sheetnames:
        raise ValueError("workbook does not contain a Review Queue sheet")
    sheet = workbook["Review Queue"]
    headers = {str(cell.value).strip(): index for index, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1)), 1) if cell.value is not None}
    required = {"Review ID", "Reviewer status", "Gold evidence chunk IDs", "Accepted graph path node IDs", "Reviewer note"}
    missing = required - set(headers)
    if missing:
        raise ValueError(f"workbook is missing required columns: {sorted(missing)}")
    decisions = {}
    for cells in sheet.iter_rows(min_row=2, values_only=True):
        annotation_id = str(cells[headers["Review ID"] - 1] or "").strip()
        if not annotation_id:
            continue
        if annotation_id in decisions:
            raise ValueError(f"duplicate Review ID in workbook: {annotation_id}")
        decisions[annotation_id] = {
            "status": str(cells[headers["Reviewer status"] - 1] or "Pending").strip(),
            "gold": str(cells[headers["Gold evidence chunk IDs"] - 1] or "").strip(),
            "path": str(cells[headers["Accepted graph path node IDs"] - 1] or "").strip(),
            "note": str(cells[headers["Reviewer note"] - 1] or "").strip(),
        }
    return decisions


def apply_decisions(pack: dict[str, Any], decisions: dict[str, dict[str, str]], chunk_ids: set[str], node_ids: set[str]) -> dict[str, Any]:
    result = deepcopy(pack)
    by_id = {str(row["annotation_id"]): row for row in result.get("queries", [])}
    unknown = sorted(set(decisions) - set(by_id))
    if unknown:
        raise ValueError(f"workbook contains unknown Review ID values: {unknown}")
    accepted = {"confirmed", "revise", "exclude", "pending"}
    changed = 0
    for annotation_id, decision in decisions.items():
        row = by_id[annotation_id]
        status = decision["status"].casefold()
        if status not in accepted:
            raise ValueError(f"unsupported reviewer status for {annotation_id}: {decision['status']!r}")
        if status == "pending":
            continue
        changed += 1
        row["reviewer_note"] = decision["note"]
        if status == "confirmed":
            gold_ids = split_ids(decision["gold"])
            unknown_chunks = sorted(set(gold_ids) - chunk_ids)
            if not gold_ids:
                raise ValueError(f"confirmed row needs at least one gold evidence chunk: {annotation_id}")
            if unknown_chunks:
                raise ValueError(f"confirmed row contains chunk IDs outside frozen corpus: {annotation_id}: {unknown_chunks}")
            path_nodes = split_path_nodes(decision["path"])
            unknown_nodes = sorted(set(path_nodes) - node_ids)
            if unknown_nodes:
                raise ValueError(f"confirmed row contains graph nodes outside frozen snapshot: {annotation_id}: {unknown_nodes}")
            row["gold_evidence_chunk_ids"] = gold_ids
            row["accepted_graph_path_node_ids"] = path_nodes
            row["review_status"] = "reviewed"
            row["eligible_for_formal_evaluation"] = True
            row["exclusion_reason"] = ""
        elif status == "revise":
            row["review_status"] = "needs_revision"
            row["eligible_for_formal_evaluation"] = False
        else:
            row["review_status"] = "excluded"
            row["eligible_for_formal_evaluation"] = False
            row["exclusion_reason"] = f"reviewer_excluded: {decision['note']}".rstrip()
    result["status"] = "reviewer_workbook_applied_not_frozen"
    result["formal_metrics_ready"] = False
    result["reviewer_workbook_application"] = {"changed_row_count": changed, "ledger": review_ledger(result["queries"])}
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply a completed three-path review workbook to a new JSON snapshot")
    parser.add_argument("--pack", required=True)
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--corpus", default=str(DEFAULT_CORPUS))
    parser.add_argument("--graph", default=str(DEFAULT_GRAPH))
    args = parser.parse_args()
    output = Path(args.output)
    if output.exists():
        raise RuntimeError(f"refusing to overwrite existing reviewed snapshot: {output}")
    pack_path = Path(args.pack)
    workbook_path = Path(args.workbook)
    result = apply_decisions(read_json(pack_path), workbook_decisions(workbook_path), frozen_chunk_ids(Path(args.corpus)), graph_node_ids(Path(args.graph)))
    result["reviewer_workbook_application"]["provenance"] = {
        "pack": str(pack_path),
        "pack_sha256": sha256_file(pack_path),
        "workbook": str(workbook_path),
        "workbook_sha256": sha256_file(workbook_path),
    }
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result["reviewer_workbook_application"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
