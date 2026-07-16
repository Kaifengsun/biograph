"""Freeze final Gold evidence after joint adjudication and activate the formal pack."""

from __future__ import annotations

import argparse
import hashlib
import json
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from tools.dual_annotation_60.analyze_dual_annotations import (
    CORE_LABELS,
    PASSAGE_LABELS,
    QUESTION_LABELS,
    COMPLETENESS_LABELS,
    parse_a_docx,
    parse_b_workbooks,
    read_json,
    write_json,
)


def sha256_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def parse_final_adjudication(
    path: Path,
    disagreements: dict[str, Any],
) -> tuple[dict[str, dict[str, str]], dict[tuple[str, str], str]]:
    workbook = load_workbook(path, data_only=False)
    if [sheet.title for sheet in workbook.worksheets] != ["操作与统计", "问题级裁决", "核心证据裁决", "标签说明"]:
        raise ValueError("unexpected final adjudication workbook structure")
    expected_questions = {row["query_id"]: row for row in disagreements["question_disagreements"]}
    expected_passages = {(row["query_id"], row["passage_id"]): row for row in disagreements["passage_disagreements"]}

    final_questions: dict[str, dict[str, str]] = {}
    sheet = workbook.worksheets[1]
    for row_index in range(5, sheet.max_row + 1):
        query_id = str(sheet.cell(row_index, 2).value or "").strip()
        if not query_id:
            continue
        expected = expected_questions.get(query_id)
        if expected is None:
            raise ValueError(f"unexpected adjudicated question: {query_id}")
        immutable = {
            "query_slice": str(sheet.cell(row_index, 3).value or "").strip(),
            "question": str(sheet.cell(row_index, 4).value or "").strip(),
            "a_status": str(sheet.cell(row_index, 5).value or "").strip(),
            "a_completeness": str(sheet.cell(row_index, 6).value or "").strip(),
            "b_status": str(sheet.cell(row_index, 8).value or "").strip(),
            "b_completeness": str(sheet.cell(row_index, 9).value or "").strip(),
        }
        expected_immutable = {
            "query_slice": expected["query_slice"], "question": expected["question"],
            "a_status": expected["a"]["question_status"],
            "a_completeness": expected["a"]["evidence_set_completeness"],
            "b_status": expected["b"]["question_status"],
            "b_completeness": expected["b"]["evidence_set_completeness"],
        }
        if immutable != expected_immutable:
            raise ValueError(f"immutable question adjudication fields changed: {query_id}")
        status = str(sheet.cell(row_index, 11).value or "").strip()
        completeness = str(sheet.cell(row_index, 12).value or "").strip()
        if status not in QUESTION_LABELS or completeness not in COMPLETENESS_LABELS:
            raise ValueError(f"invalid final question decision: {query_id}: {status}/{completeness}")
        final_questions[query_id] = {"question_status": status, "evidence_set_completeness": completeness}
    if set(final_questions) != set(expected_questions):
        raise ValueError("final question adjudication rows do not match disagreement registry")

    final_passages: dict[tuple[str, str], str] = {}
    sheet = workbook.worksheets[2]
    for row_index in range(5, sheet.max_row + 1):
        query_id = str(sheet.cell(row_index, 2).value or "").strip()
        passage_id = str(sheet.cell(row_index, 5).value or "").strip()
        if not query_id or not passage_id:
            continue
        key = (query_id, passage_id)
        expected = expected_passages.get(key)
        if expected is None:
            raise ValueError(f"unexpected adjudicated passage: {key}")
        immutable = {
            "query_slice": str(sheet.cell(row_index, 3).value or "").strip(),
            "question": str(sheet.cell(row_index, 4).value or "").strip(),
            "source_document": str(sheet.cell(row_index, 6).value or "").strip(),
            "heading": str(sheet.cell(row_index, 7).value or "").strip(),
            "content": str(sheet.cell(row_index, 8).value or "").strip(),
            "a_label": str(sheet.cell(row_index, 9).value or "").strip(),
            "b_label": str(sheet.cell(row_index, 11).value or "").strip(),
            "reason": str(sheet.cell(row_index, 13).value or "").strip(),
        }
        expected_immutable = {
            "query_slice": expected["query_slice"], "question": expected["question"],
            "source_document": expected["source_document"], "heading": expected["heading"],
            "content": expected["frozen_source_passage"],
            "a_label": expected["a_label"] or "[Missing]",
            "b_label": expected["b_label"] or "[Missing]", "reason": expected["reason"],
        }
        if immutable != expected_immutable:
            raise ValueError(f"immutable passage adjudication fields changed: {key}")
        label = str(sheet.cell(row_index, 14).value or "").strip()
        if label not in PASSAGE_LABELS:
            raise ValueError(f"invalid final passage decision: {key}: {label}")
        final_passages[key] = label
    if set(final_passages) != set(expected_passages):
        raise ValueError("final passage adjudication rows do not match disagreement registry")
    return final_questions, final_passages


def verify_method_lock(lock_path: Path) -> dict[str, Any]:
    lock = read_json(lock_path)
    if lock.get("lock_status") != "locked_before_confirmatory_set_construction":
        raise ValueError("method lock has unexpected status")
    changed = []
    for raw_path, expected_hash in lock.get("inputs", {}).items():
        path = ROOT / raw_path
        if not path.exists() or sha256_file(path) != expected_hash:
            changed.append(raw_path)
    if changed:
        raise ValueError(f"locked method inputs changed after lock: {changed[:10]}")
    return lock


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--registry", type=Path, required=True)
    parser.add_argument("--reviewer-a", type=Path, required=True)
    parser.add_argument("--reviewer-b", type=Path, nargs=3, required=True)
    parser.add_argument("--disagreements", type=Path, required=True)
    parser.add_argument("--final-adjudication", type=Path, required=True)
    parser.add_argument("--method-lock", type=Path, required=True)
    parser.add_argument("--full-output", type=Path, required=True)
    parser.add_argument("--formal-output", type=Path, required=True)
    args = parser.parse_args()
    if args.full_output.exists() or args.formal_output.exists():
        raise RuntimeError("refusing to overwrite frozen final annotation outputs")

    registry = read_json(args.registry)
    disagreements = read_json(args.disagreements)
    query_lookup = {row["query_id"]: row for row in registry["queries"]}
    question_a, passage_a = parse_a_docx(args.reviewer_a)
    question_b, passage_b = parse_b_workbooks(args.reviewer_b)
    final_question_decisions, final_passage_decisions = parse_final_adjudication(args.final_adjudication, disagreements)
    lock = verify_method_lock(args.method_lock)

    finalized_queries = []
    for query_id, query in query_lookup.items():
        a_question, b_question = question_a[query_id], question_b[query_id]
        if query_id in final_question_decisions:
            final_question = final_question_decisions[query_id]
            question_resolution = "joint_adjudication"
        else:
            if (a_question["question_status"], a_question["evidence_set_completeness"]) != (b_question["question_status"], b_question["evidence_set_completeness"]):
                raise ValueError(f"unadjudicated question disagreement: {query_id}")
            final_question = {
                "question_status": a_question["question_status"],
                "evidence_set_completeness": a_question["evidence_set_completeness"],
            }
            question_resolution = "exact_consensus"

        passage_ledger = []
        gold_chunk_ids = []
        gold_passage_ids = []
        for passage in query["candidate_passages"]:
            passage_id = passage["blind_passage_id"]
            key = (query_id, passage_id)
            a_label = passage_a.get(key, {}).get("passage_label", "")
            b_label = passage_b.get(key, {}).get("passage_label", "")
            if key in final_passage_decisions:
                final_label = final_passage_decisions[key]
                final_core_gold = final_label in CORE_LABELS
                resolution = "joint_adjudication"
            else:
                if not a_label or not b_label:
                    raise ValueError(f"missing unadjudicated passage label: {key}")
                if (a_label in CORE_LABELS) != (b_label in CORE_LABELS):
                    raise ValueError(f"unadjudicated core Gold disagreement: {key}")
                final_core_gold = a_label in CORE_LABELS
                final_label = a_label if a_label == b_label else None
                resolution = "exact_consensus" if a_label == b_label else "core_binary_consensus"
            if final_core_gold:
                gold_chunk_ids.append(passage["chunk_id"])
                gold_passage_ids.append(passage_id)
            passage_ledger.append({
                "blind_passage_id": passage_id,
                "chunk_id": passage["chunk_id"],
                "reviewer_a_label": a_label or None,
                "reviewer_b_label": b_label or None,
                "final_passage_label": final_label,
                "final_core_gold": final_core_gold,
                "resolution": resolution,
            })

        eligible = (
            final_question["question_status"] == "Answerable"
            and final_question["evidence_set_completeness"] == "Complete"
            and bool(gold_chunk_ids)
        )
        finalized_queries.append({
            "annotation_id": query_id,
            "batch": query["batch"],
            "query_slice": query["query_slice"],
            "query": query["query"],
            "final_question_status": final_question["question_status"],
            "final_evidence_set_completeness": final_question["evidence_set_completeness"],
            "question_resolution": question_resolution,
            "eligible_for_formal_evaluation": eligible,
            "gold_evidence_chunk_ids": list(dict.fromkeys(gold_chunk_ids)),
            "gold_blind_passage_ids": gold_passage_ids,
            "passage_ledger": passage_ledger,
        })

    eligible_queries = [row for row in finalized_queries if row["eligible_for_formal_evaluation"]]
    if len(finalized_queries) != 60 or len(eligible_queries) != 58:
        raise ValueError(f"expected 60 final and 58 eligible queries, got {len(finalized_queries)}/{len(eligible_queries)}")
    if any(not row["gold_evidence_chunk_ids"] for row in eligible_queries):
        raise ValueError("eligible query has no final Gold evidence")

    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    hashes = {
        "initial_registry": sha256_file(args.registry),
        "reviewer_a": sha256_file(args.reviewer_a),
        "reviewer_b": [sha256_file(path) for path in args.reviewer_b],
        "disagreement_registry": sha256_file(args.disagreements),
        "final_adjudication": sha256_file(args.final_adjudication),
        "method_lock": sha256_file(args.method_lock),
    }
    full_payload = {
        "schema_version": "1.0",
        "status": "frozen_jointly_adjudicated_complete",
        "finalized_at_utc": now,
        "total_queries": 60,
        "eligible_queries": 58,
        "excluded_queries": [row["annotation_id"] for row in finalized_queries if not row["eligible_for_formal_evaluation"]],
        "no_unresolved_decisions": True,
        "input_hashes": hashes,
        "queries": finalized_queries,
    }
    formal_queries = [{
        key: deepcopy(row[key]) for key in (
            "annotation_id", "batch", "query_slice", "query", "gold_evidence_chunk_ids"
        )
    } for row in eligible_queries]
    formal_payload = {
        "schema_version": "1.0",
        "status": "frozen_jointly_adjudicated_formal_run_ready",
        "formal_metrics_ready": True,
        "retrieval_execution_prohibited": False,
        "confirmatory_for_source_chunk_reranker": True,
        "single_formal_execution_authorized": True,
        "created_at_utc": now,
        "query_count": len(formal_queries),
        "excluded_query_ids": full_payload["excluded_queries"],
        "method_lock": {"path": str(args.method_lock), "sha256": hashes["method_lock"], "method_family": lock["method_family"]},
        "input_hashes": hashes,
        "queries": formal_queries,
    }
    write_json(args.full_output, full_payload)
    write_json(args.formal_output, formal_payload)
    print(json.dumps({
        "full_output": str(args.full_output),
        "formal_output": str(args.formal_output),
        "eligible_queries": len(formal_queries),
        "excluded_query_ids": full_payload["excluded_queries"],
        "no_unresolved_decisions": True,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
